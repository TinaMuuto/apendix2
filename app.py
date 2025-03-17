import streamlit as st
import pandas as pd
from io import BytesIO
from openpyxl import load_workbook

@st.cache_data
def load_template():
    """Læser template-filen (apendix2-template.xlsx) lokalt."""
    with open("apendix2-template.xlsx", "rb") as f:
        return f.read()

@st.cache_data
def load_masterdata():
    """Læser masterdata-filen (Muuto_Master_Data_CON_January_2025_DKK.xlsx) lokalt."""
    return pd.read_excel("Muuto_Master_Data_CON_January_2025_DKK.xlsx", engine="openpyxl")

@st.cache_data
def load_family():
    """Læser family-filen (family.xlsx) lokalt."""
    return pd.read_excel("family.xlsx", engine="openpyxl")

st.title("App til generereing af EY appendix 2")
st.write("Indtast varenumre (ét varenummer per linje):")
user_input = st.text_area("Varenumre", height=200)

if st.button("Start behandling"):
    # Simpel status og progress
    progress_bar = st.progress(0)
    status_text = st.empty()
    status_text.text("Behandler...")

    # Hent filer
    template_content = load_template()
    masterdata_df = load_masterdata()
    family_df = load_family()

    # Del brugerens input i en liste (fjerner evt. tomme linjer)
    varenumre = [line.strip() for line in user_input.splitlines() if line.strip()]
    results = []
    unmatched = []

    total = len(varenumre)

    for i, varenr in enumerate(varenumre):
        # 1) Opslag i masterdata baseret på "ITEM NO."
        match = masterdata_df[masterdata_df["ITEM NO."].astype(str) == varenr]
        
        # Hvis ingen eksakt match, forsøg delvist match (split ved " - ")
        if match.empty:
            partial_matches = masterdata_df[
                masterdata_df["ITEM NO."].astype(str).apply(
                    lambda x: x.split(" - ")[0] == varenr
                )
            ]
            if not partial_matches.empty:
                match = partial_matches.iloc[[0]]

        if not match.empty:
            row = match.iloc[0]
            # Nu slår vi op i family-filen for at få "Product Series name"
            # (samme logik: eksakt match først, ellers delvist match)
            match_family = family_df[family_df["ITEM NO."].astype(str) == varenr]

            if match_family.empty:
                partial_fam = family_df[
                    family_df["ITEM NO."].astype(str).apply(
                        lambda x: x.split(" - ")[0] == varenr
                    )
                ]
                if not partial_fam.empty:
                    match_family = partial_fam.iloc[[0]]

            if not match_family.empty:
                row_family = match_family.iloc[0]
                product_series_name = row_family["Family"]  # Fra kolonnen "Family"
            else:
                # Hvis intet match i family-filen, sættes den til tom
                product_series_name = ""

            # Hent data fra masterdata
            product_name = row["PRODUCT"]  # Tilpas hvis kolonnenavn er anderledes
            product_item_number = varenr
            product_description = row["PRODUCT DESCRIPTION"]
            manufacturing_country = row["COUNTRY OF ORIGIN"]
            lead_time = row["LEAD TIME"]
            if str(lead_time).strip() == "-":
                lead_time = "Ready to ship"
            warranty = row["WARRANTY"]
            contract_price = row["CONTRACT PRICE"]
            list_price = f"{contract_price} DKK"

            results.append({
                "Product Series name": product_series_name,
                "Product Name": product_name,
                "Product item number": product_item_number,
                "Product Description": product_description,
                "Manufacturing country": manufacturing_country,
                "Lead time [weeks]": lead_time,
                "Product Guarantee period [years]": warranty,
                "List Price [your currency]": list_price
            })
        else:
            # Ingen match i masterdata
            unmatched.append(varenr)

        progress_bar.progress((i+1) / total)

    # Hvis vi ikke fik nogle resultater
    if not results:
        st.error("Ingen matches fundet i masterdata. Tjek dine varenumre.")
    else:
        # Åbn template-filen med openpyxl
        wb = load_workbook(filename=BytesIO(template_content))
        ws = wb.active

        # Antag at kolonneoverskrifter ligger i B7 til I7 => data starter i række 8
        start_row = 8
        for idx, res in enumerate(results, start=start_row):
            ws[f"B{idx}"] = res["Product Series name"]
            ws[f"C{idx}"] = res["Product Name"]
            ws[f"D{idx}"] = res["Product item number"]
            ws[f"E{idx}"] = res["Product Description"]
            ws[f"F{idx}"] = res["Manufacturing country"]
            ws[f"G{idx}"] = res["Lead time [weeks]"]
            ws[f"H{idx}"] = res["Product Guarantee period [years]"]
            ws[f"I{idx}"] = res["List Price [your currency]"]

        # Gem den udfyldte fil i en BytesIO-strøm
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        st.success("Behandlingen er færdig!")
        st.download_button(
            "Download udfyldt fil",
            data=output,
            file_name="udfyldt_template.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    if unmatched:
        st.warning("Følgende varenumre blev ikke fundet i masterdata:")
        st.text("\n".join(unmatched))
